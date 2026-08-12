import os
import re
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def get_time_category(time_str):
    try:
        clean_time = time_str.strip()
        # Convert "8 o'clock PM" to "8:00 PM"
        clean_time = re.sub(r"(?i)\s*o'clock\s*", ":00 ", clean_time)
        t = datetime.strptime(clean_time.strip(), "%I:%M %p")
        h = t.hour
        if 5 <= h < 9: return "1. Early Morning (5 AM - 9 AM)"
        elif 9 <= h < 12: return "2. Morning (9 AM - 12 PM)"
        elif 12 <= h < 16: return "3. Afternoon (12 PM - 4 PM)"
        elif 16 <= h < 20: return "4. Evening (4 PM - 8 PM)"
        elif 20 <= h < 24: return "5. Night (8 PM - 12 AM)"
        else: return "6. Midnight (12 AM - 5 AM)"
    except:
        return "7. Unknown Time"

def get_chain_category(theater_name):
    name = theater_name.upper()
    if "AMC" in name: return "AMC Theatres"
    if "CINEMARK" in name or "CENTURY" in name: return "Cinemark"
    if "REGAL" in name: return "Regal Cinemas"
    if "MARCUS" in name: return "Marcus Theatres"
    if "HARKINS" in name: return "Harkins Theatres"
    if "APPLE CINEMAS" in name: return "Apple Cinemas"
    return "Other / Independents"

def create_chain_summary(data, ws, p_shows_map):
    summary = {}
    for row in data:
        if row.get('is_extra', False): continue
        chain = get_chain_category(row['theater'])
        if chain not in summary:
            summary[chain] = {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0}
        
        summary[chain]['shows'] += 1
        summary[chain]['tickets'] += row['total']
        summary[chain]['booked'] += row['booked']
        summary[chain]['gross'] += row['gross']

        fmt = row['format']
        if 'D-Box' in fmt and 'Premium' in fmt: fmt = 'Premium'
        show_key = f"{row['t_id']}_{row['time']}_{fmt}"
        p_show = p_shows_map.get(show_key)
        
        if not p_show:
            summary[chain]['d_shows'] += 1
            summary[chain]['d_tickets'] += row['booked']
            summary[chain]['d_gross'] += row['gross']
        else:
            summary[chain]['d_tickets'] += (row['booked'] - p_show['booked'])
            summary[chain]['d_gross'] += (row['gross'] - p_show['gross'])

    headers = ['Theater Chain', 'Total Shows', 'Tickets', 'Tickets Sold', 'Gross ($)', 'Occupancy %', 'Shows Growth', 'Tickets Growth', 'Gross Growth']
    ws.append(headers)

    for chain, stats in sorted(summary.items(), key=lambda x: x[1]['gross'], reverse=True):
        occ = round((stats['booked'] / stats['tickets']) * 100, 2) if stats['tickets'] > 0 else 0
        ws.append([
            chain, stats['shows'], stats['tickets'], stats['booked'], 
            round(stats['gross'], 2), occ, stats['d_shows'], 
            stats['d_tickets'], round(stats['d_gross'], 2)
        ])

def create_language_summary(data, ws, p_shows_map):
    summary = {}
    for row in data:
        if row.get('is_extra', False): continue
        lang = row.get('language', 'Unknown')
        if lang not in summary:
            summary[lang] = {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0}
        
        summary[lang]['shows'] += 1
        summary[lang]['tickets'] += row['total']
        summary[lang]['booked'] += row['booked']
        summary[lang]['gross'] += row['gross']
        
        fmt = row['format']
        if 'D-Box' in fmt and 'Premium' in fmt: fmt = 'Premium'
        show_key = f"{row['t_id']}_{row['time']}_{fmt}"
        p_show = p_shows_map.get(show_key)
        
        if not p_show:
            summary[lang]['d_shows'] += 1
            summary[lang]['d_tickets'] += row['booked']
            summary[lang]['d_gross'] += row['gross']
        else:
            summary[lang]['d_tickets'] += (row['booked'] - p_show['booked'])
            summary[lang]['d_gross'] += (row['gross'] - p_show['gross'])

    headers = ['Language', 'Total Shows', 'Tickets', 'Tickets Sold', 'Gross ($)', 'Occupancy %', 'Shows Growth', 'Tickets Growth', 'Gross Growth']
    ws.append(headers)

    for lang, stats in sorted(summary.items(), key=lambda x: x[1]['gross'], reverse=True):
        occ = round((stats['booked'] / stats['tickets']) * 100, 2) if stats['tickets'] > 0 else 0
        ws.append([
            lang, stats['shows'], stats['tickets'], stats['booked'], 
            round(stats['gross'], 2), occ, stats['d_shows'], 
            stats['d_tickets'], round(stats['d_gross'], 2)
        ])

def export_master_excel(data, base_filename, previous_shows_data=None, last_updated_str="N/A"):
    wb = Workbook()
    
    # Hijack the default "Sheet" so we don't have an empty tab at the end
    ws_kpi = wb.active
    ws_kpi.title = "Summary KPIs"
    
    # Create the rest of the sheets
    ws_chain = wb.create_sheet("Chain Summary")
    ws_lang = wb.create_sheet("Language Summary")
    ws_shows = wb.create_sheet("Showtime Details")
    ws_theaters = wb.create_sheet("Theater Summary")
    ws_states = wb.create_sheet("State Summary")
    ws_occ = wb.create_sheet("Occupancy Tiers")
    ws_time = wb.create_sheet("Time of Day")
    ws_fmt = wb.create_sheet("Format Summary")

    # Convert previous data into a map for fast lookup
    p_shows_map = {}
    if previous_shows_data:
        for r in previous_shows_data:
            fmt = r.get('format', '')
            if 'D-Box' in fmt and 'Premium' in fmt: fmt = 'Premium'
            key = f"{r.get('t_id')}_{r.get('time')}_{fmt}"
            p_shows_map[key] = r

    # =========================================================================
    # 1. Shows Sheet
    # =========================================================================
    headers = ['State', 'Chain', 'Format', 'Language', 'Theater Name', 'Show Time', 'Time Category', 'Status', 'Ticket Price', 'Tickets', 'Booked', 'Gross ($)', 'Occupancy %', 'Tickets Growth', 'Gross Growth']
    ws_shows.append(headers)

    for row in data:
        if row.get('is_extra', False): continue
            
        occ = round((row['booked'] / row['total']) * 100, 2) if row['total'] > 0 else 0
        fmt = row['format']
        if 'D-Box' in fmt and 'Premium' in fmt: fmt = 'Premium'
        chain = get_chain_category(row['theater'])
        time_cat = get_time_category(row['time'])
        
        # Calculate Delta
        show_key = f"{row['t_id']}_{row['time']}_{fmt}"
        p_show = p_shows_map.get(show_key, {'booked': 0, 'gross': 0.0})
        
        d_booked = row['booked'] - p_show['booked']
        d_gross = row['gross'] - p_show['gross']
            
        ws_shows.append([
            row['state'], 
            chain, 
            fmt, 
            row.get('language', 'Unknown'),
            row['theater'], 
            row['time'], 
            time_cat, 
            row.get('status', 'Available'), 
            row.get('price_str', '$0.00'), 
            row['total'], 
            row['booked'], 
            round(row['gross'], 2), 
            occ,
            d_booked,
            round(d_gross, 2)
        ])

    # =========================================================================
    # Aggregations
    # =========================================================================
    
    create_chain_summary(data, ws_chain, p_shows_map)
    create_language_summary(data, ws_lang, p_shows_map)

    # 2. Theaters Sheet
    t_summary = {}
    for row in data:
        t_id = row['t_id']
        if t_id == 'EXTRA': continue
        
        if t_id not in t_summary:
            t_summary[t_id] = {'state': row['state'], 'name': row['theater'], 'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0}
        t_summary[t_id]['shows'] += 1
        t_summary[t_id]['tickets'] += row['total']
        t_summary[t_id]['booked'] += row['booked']
        t_summary[t_id]['gross'] += row['gross']
        
        fmt = row['format']
        if 'D-Box' in fmt and 'Premium' in fmt: fmt = 'Premium'
        show_key = f"{row['t_id']}_{row['time']}_{fmt}"
        p_show = p_shows_map.get(show_key)
        
        if not p_show:
            t_summary[t_id]['d_shows'] += 1
            t_summary[t_id]['d_tickets'] += row['booked']
            t_summary[t_id]['d_gross'] += row['gross']
        else:
            t_summary[t_id]['d_tickets'] += (row['booked'] - p_show['booked'])
            t_summary[t_id]['d_gross'] += (row['gross'] - p_show['gross'])

    ws_theaters.append(['State', 'Theater Name', 'Total Shows', 'Tickets', 'Tickets Sold', 'Gross ($)', 'Occupancy %', 'Shows Growth', 'Tickets Growth', 'Gross Growth'])
    for t_id, stats in sorted(t_summary.items(), key=lambda x: x[1]['gross'], reverse=True):
        occ = round((stats['booked'] / stats['tickets']) * 100, 2) if stats['tickets'] > 0 else 0
        ws_theaters.append([stats['state'], stats['name'], stats['shows'], stats['tickets'], stats['booked'], round(stats['gross'], 2), occ, stats['d_shows'], stats['d_tickets'], round(stats['d_gross'], 2)])

    # 3. States Sheet
    s_summary = {}
    for row in data:
        st = row['state']
        if st == 'EXTRA': continue
        
        if st not in s_summary:
            s_summary[st] = {'theaters': set(), 'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0}
        s_summary[st]['theaters'].add(row['t_id'])
        s_summary[st]['shows'] += 1
        s_summary[st]['tickets'] += row['total']
        s_summary[st]['booked'] += row['booked']
        s_summary[st]['gross'] += row['gross']
        
        fmt = row['format']
        if 'D-Box' in fmt and 'Premium' in fmt: fmt = 'Premium'
        show_key = f"{row['t_id']}_{row['time']}_{fmt}"
        p_show = p_shows_map.get(show_key)
        
        if not p_show:
            s_summary[st]['d_shows'] += 1
            s_summary[st]['d_tickets'] += row['booked']
            s_summary[st]['d_gross'] += row['gross']
        else:
            s_summary[st]['d_tickets'] += (row['booked'] - p_show['booked'])
            s_summary[st]['d_gross'] += (row['gross'] - p_show['gross'])

    ws_states.append(['State', 'Total Theaters', 'Total Shows', 'Tickets', 'Tickets Sold', 'Gross ($)', 'Occupancy %', 'Shows Growth', 'Tickets Growth', 'Gross Growth'])
    for st, stats in sorted(s_summary.items(), key=lambda x: x[1]['gross'], reverse=True):
        occ = round((stats['booked'] / stats['tickets']) * 100, 2) if stats['tickets'] > 0 else 0
        ws_states.append([st, len(stats['theaters']), stats['shows'], stats['tickets'], stats['booked'], round(stats['gross'], 2), occ, stats['d_shows'], stats['d_tickets'], round(stats['d_gross'], 2)])

    # 4. Occupancy Tiers
    occ_summary = {
        '01. Empty (0%)': {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0},
        '02. Low (<30%)': {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0},
        '03. Medium (30-60%)': {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0},
        '04. High (60-90%)': {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0},
        '05. Sold Out / Near Full (>90%)': {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0},
        'Unknown': {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0}
    }
    for row in data:
        if row.get('is_extra', False): continue
            
        occ = (row['booked'] / row['total'] * 100) if row['total'] > 0 else 0
        if row['total'] == 0: cat = 'Unknown'
        elif occ == 0: cat = '01. Empty (0%)'
        elif occ < 30: cat = '02. Low (<30%)'
        elif occ < 60: cat = '03. Medium (30-60%)'
        elif occ < 90: cat = '04. High (60-90%)'
        else: cat = '05. Sold Out / Near Full (>90%)'
        
        occ_summary[cat]['shows'] += 1
        occ_summary[cat]['tickets'] += row['total']
        occ_summary[cat]['booked'] += row['booked']
        occ_summary[cat]['gross'] += row['gross']
        
        fmt = row['format']
        if 'D-Box' in fmt and 'Premium' in fmt: fmt = 'Premium'
        show_key = f"{row['t_id']}_{row['time']}_{fmt}"
        p_show = p_shows_map.get(show_key)
        
        if not p_show:
            occ_summary[cat]['d_shows'] += 1
            occ_summary[cat]['d_tickets'] += row['booked']
            occ_summary[cat]['d_gross'] += row['gross']
        else:
            occ_summary[cat]['d_tickets'] += (row['booked'] - p_show['booked'])
            occ_summary[cat]['d_gross'] += (row['gross'] - p_show['gross'])

    ws_occ.append(['Occupancy Level', 'Total Shows', 'Tickets', 'Tickets Sold', 'Gross ($)', 'Occupancy %', 'Shows Growth', 'Tickets Growth', 'Gross Growth'])
    for cat in sorted(occ_summary.keys()):
        stats = occ_summary[cat]
        real_occ = round((stats['booked'] / stats['tickets']) * 100, 2) if stats['tickets'] > 0 else 0
        ws_occ.append([cat, stats['shows'], stats['tickets'], stats['booked'], round(stats['gross'], 2), real_occ, stats['d_shows'], stats['d_tickets'], round(stats['d_gross'], 2)])

    # 5. Time of Day
    time_summary = {}
    for row in data:
        if row.get('is_extra', False): continue
            
        t_cat = get_time_category(row['time'])
        if t_cat not in time_summary:
            time_summary[t_cat] = {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0}
        time_summary[t_cat]['shows'] += 1
        time_summary[t_cat]['tickets'] += row['total']
        time_summary[t_cat]['booked'] += row['booked']
        time_summary[t_cat]['gross'] += row['gross']
        
        fmt = row['format']
        if 'D-Box' in fmt and 'Premium' in fmt: fmt = 'Premium'
        show_key = f"{row['t_id']}_{row['time']}_{fmt}"
        p_show = p_shows_map.get(show_key)
        
        if not p_show:
            time_summary[t_cat]['d_shows'] += 1
            time_summary[t_cat]['d_tickets'] += row['booked']
            time_summary[t_cat]['d_gross'] += row['gross']
        else:
            time_summary[t_cat]['d_tickets'] += (row['booked'] - p_show['booked'])
            time_summary[t_cat]['d_gross'] += (row['gross'] - p_show['gross'])

    ws_time.append(['Time Category', 'Total Shows', 'Tickets', 'Tickets Sold', 'Gross ($)', 'Occupancy %', 'Shows Growth', 'Tickets Growth', 'Gross Growth'])
    for cat in sorted(time_summary.keys()):
        stats = time_summary[cat]
        occ = round((stats['booked'] / stats['tickets']) * 100, 2) if stats['tickets'] > 0 else 0
        ws_time.append([cat, stats['shows'], stats['tickets'], stats['booked'], round(stats['gross'], 2), occ, stats['d_shows'], stats['d_tickets'], round(stats['d_gross'], 2)])

    # 6. Format Summary
    fmt_summary = {}
    for row in data:
        if row.get('is_extra', False): continue
            
        fmt = row['format']
        if 'D-Box' in fmt and 'Premium' in fmt: fmt = 'Premium'
            
        if fmt not in fmt_summary:
            fmt_summary[fmt] = {'shows': 0, 'tickets': 0, 'booked': 0, 'gross': 0.0, 'd_shows': 0, 'd_tickets': 0, 'd_gross': 0.0}
        fmt_summary[fmt]['shows'] += 1
        fmt_summary[fmt]['tickets'] += row['total']
        fmt_summary[fmt]['booked'] += row['booked']
        fmt_summary[fmt]['gross'] += row['gross']
        
        show_key = f"{row['t_id']}_{row['time']}_{fmt}"
        p_show = p_shows_map.get(show_key)
        
        if not p_show:
            fmt_summary[fmt]['d_shows'] += 1
            fmt_summary[fmt]['d_tickets'] += row['booked']
            fmt_summary[fmt]['d_gross'] += row['gross']
        else:
            fmt_summary[fmt]['d_tickets'] += (row['booked'] - p_show['booked'])
            fmt_summary[fmt]['d_gross'] += (row['gross'] - p_show['gross'])

    ws_fmt.append(['Screen Format', 'Total Shows', 'Tickets', 'Tickets Sold', 'Gross ($)', 'Occupancy %', 'Shows Growth', 'Tickets Growth', 'Gross Growth'])
    for fmt, stats in sorted(fmt_summary.items(), key=lambda x: x[1]['gross'], reverse=True):
        occ = round((stats['booked'] / stats['tickets']) * 100, 2) if stats['tickets'] > 0 else 0
        ws_fmt.append([fmt, stats['shows'], stats['tickets'], stats['booked'], round(stats['gross'], 2), occ, stats['d_shows'], stats['d_tickets'], round(stats['d_gross'], 2)])

    # 7. Overall KPIs (Top Sheet)
    total_venues = len(set(r['t_id'] for r in data if r['t_id'] != 'EXTRA'))
    total_shows = sum(1 for r in data if not r.get('is_extra', False))
    total_tickets = sum(r['total'] for r in data)
    total_booked = sum(r['booked'] for r in data)
    total_gross = sum(r['gross'] for r in data)
    overall_occ = round((total_booked / total_tickets) * 100, 2) if total_tickets > 0 else 0
    atp = round(total_gross / total_booked, 2) if total_booked > 0 else 0

    p_total_venues = len(set(r['t_id'] for r in previous_shows_data if r['t_id'] != 'EXTRA')) if previous_shows_data else 0
    p_total_shows = sum(1 for r in previous_shows_data if not r.get('is_extra', False)) if previous_shows_data else 0
    p_total_booked = sum(r['booked'] for r in previous_shows_data) if previous_shows_data else 0
    p_total_gross = sum(r['gross'] for r in previous_shows_data) if previous_shows_data else 0

    d_venues = total_venues - p_total_venues
    d_shows = total_shows - p_total_shows
    d_booked = total_booked - p_total_booked
    d_gross = total_gross - p_total_gross

    def format_delta(val, is_currency=False):
        if previous_shows_data is None: return "-"
        if val == 0: return "-"
        sign = "+" if val > 0 else ""
        if is_currency:
            return f"{sign}${val:,.2f}"
        return f"{sign}{val:,}"

    ws_kpi.append(['Metric', 'Value', f'Growth (Since {last_updated_str})'])
    
    kpis = [
        ['Total Venues', total_venues, format_delta(d_venues)],
        ['Total Shows', total_shows, format_delta(d_shows)],
        ['Total Tickets', total_tickets, "-"],
        ['Total Tickets Sold', total_booked, format_delta(d_booked)],
        ['Overall Occupancy (%)', overall_occ, "-"],
        ['Total Gross ($)', round(total_gross, 2), format_delta(d_gross, True)],
        ['Average Ticket Price (ATP) ($)', atp, "-"]
    ]
    for kpi in kpis:
        ws_kpi.append(kpi)

    # Style KPIs
    for row in ws_kpi.iter_rows(min_row=1, max_row=8, min_col=1, max_col=3):
        for cell in row:
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            else:
                cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal='left')

    ws_kpi.column_dimensions['A'].width = 30
    ws_kpi.column_dimensions['B'].width = 15
    ws_kpi.column_dimensions['C'].width = 35

    # Auto-adjust column widths for all sheets
    for ws in wb.worksheets:
        if ws.title == "Summary KPIs": continue
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 40)

    # =========================================================================
    # APPLY HEADER FORMATTING SAFELY (Fixes the Empty Row 1 Bug)
    # =========================================================================
    for ws in wb.worksheets:
        if ws.title == "Summary KPIs": 
            continue # Skip KPI tab so we don't overwrite its custom blue styling
        for cell in ws[1]:
            cell.font = Font(bold=True)

    # Save
    excel_path = f"{base_filename}.xlsx"
    wb.save(excel_path)
    print(f"📊 Excel report saved to {excel_path}")
    return excel_path